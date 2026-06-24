"""测试报告导出工具 - 支持 Excel 和 PDF 格式"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


async def export_report_excel(report: dict) -> bytes:
    """将报告导出为 Excel 格式"""
    wb = Workbook()
    ws = wb.active
    ws.title = "测试报告"

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1a9f82", end_color="1a9f82", fill_type="solid")
    sub_header_fill = PatternFill(start_color="e8f5f1", end_color="e8f5f1", fill_type="solid")
    pass_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    fail_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # 报告概览
    ws['A1'] = "测试报告详情"
    ws['A1'].font = Font(bold=True, size=16, color="1a9f82")
    ws.merge_cells('A1:F1')

    # 基本信息
    ws['A3'] = "报告ID"
    ws['B3'] = report.get('id', '')
    ws['A4'] = "场景名称"
    ws['B4'] = report.get('scene_name', '未命名场景')
    ws['A5'] = "执行环境"
    ws['B5'] = report.get('env_name', '测试环境')
    ws['A6'] = "执行时间"
    ws['B6'] = report.get('created_at', '')
    ws['A7'] = "总耗时"
    ws['B7'] = f"{report.get('duration', 0):.2f}s"

    for row in range(3, 8):
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = sub_header_fill
        ws[f'B{row}'].border = thin_border

    # 统计结果
    ws['D3'] = "通过"
    ws['E3'] = report.get('pass_count', 0)
    ws['D4'] = "失败"
    ws['E4'] = report.get('fail_count', 0)
    ws['D5'] = "跳过"
    ws['E5'] = report.get('skip_count', 0)
    ws['D6'] = "总计"
    ws['E6'] = report.get('total_count', 0)
    ws['D7'] = "通过率"
    rate = (report.get('pass_count', 0) / report.get('total_count', 1) * 100) if report.get('total_count', 0) > 0 else 0
    ws['E7'] = f"{rate:.1f}%"

    for row in range(3, 8):
        ws[f'D{row}'].font = Font(bold=True)
        ws[f'D{row}'].fill = sub_header_fill
        ws[f'E{row}'].border = thin_border
        ws[f'E{row}'].alignment = center_align

    # 步骤明细表
    ws['A10'] = "步骤明细"
    ws['A10'].font = Font(bold=True, size=14, color="1a9f82")
    ws.merge_cells('A10:G10')

    headers = ["序号", "步骤名称", "方法", "URL", "状态", "耗时", "错误信息"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=11, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    steps = report.get('steps', [])
    for idx, step in enumerate(steps):
        row_num = 12 + idx
        status = step.get('status', '')
        status_text = '通过' if status == 'success' else '失败' if status in ('failed', 'error') else '跳过' if status == 'skipped' else '进行中'

        ws.cell(row=row_num, column=1, value=idx + 1).border = thin_border
        ws.cell(row=row_num, column=2, value=step.get('label', '未命名')).border = thin_border
        ws.cell(row=row_num, column=3, value=step.get('request_method', '')).border = thin_border
        ws.cell(row=row_num, column=4, value=step.get('request_url', '')).border = thin_border

        status_cell = ws.cell(row=row_num, column=5, value=status_text)
        status_cell.border = thin_border
        status_cell.alignment = center_align
        if status == 'success':
            status_cell.fill = pass_fill
        elif status in ('failed', 'error'):
            status_cell.fill = fail_fill

        ws.cell(row=row_num, column=6, value=f"{step.get('duration', 0):.3f}s").border = thin_border
        ws.cell(row=row_num, column=7, value=step.get('error_message', '') or '').border = thin_border

    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 30

    # 导出
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


async def export_report_pdf(report: dict) -> bytes:
    """将报告导出为 PDF 格式"""
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm)
    elements = []
    styles = getSampleStyleSheet()

    # 标题
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1a9f82'))
    elements.append(Paragraph("API Pilot 测试报告", title_style))
    elements.append(Spacer(1, 10*mm))

    # 基本信息表格
    info_data = [
        ["报告ID", str(report.get('id', '')), "场景名称", report.get('scene_name', '未命名场景')],
        ["执行环境", report.get('env_name', '测试环境'), "执行时间", report.get('created_at', '')],
        ["总耗时", f"{report.get('duration', 0):.2f}s", "通过率", f"{(report.get('pass_count', 0) / max(report.get('total_count', 1), 1) * 100):.1f}%"],
    ]
    info_table = Table(info_data, colWidths=[30*mm, 50*mm, 30*mm, 50*mm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8f5f1')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#e8f5f1')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 10*mm))

    # 统计结果
    stats_data = [
        ["通过", "失败", "跳过", "总计"],
        [str(report.get('pass_count', 0)), str(report.get('fail_count', 0)),
         str(report.get('skip_count', 0)), str(report.get('total_count', 0))]
    ]
    stats_table = Table(stats_data, colWidths=[40*mm, 40*mm, 40*mm, 40*mm])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a9f82')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#1a9f82')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 15*mm))

    # 步骤明细
    elements.append(Paragraph("步骤明细", styles['Heading2']))
    elements.append(Spacer(1, 5*mm))

    step_headers = ["序号", "步骤名称", "方法", "URL", "状态", "耗时"]
    steps = report.get('steps', [])
    step_data = [step_headers]
    for idx, step in enumerate(steps):
        status = step.get('status', '')
        status_text = '通过' if status == 'success' else '失败' if status in ('failed', 'error') else '跳过' if status == 'skipped' else '进行中'
        step_data.append([
            str(idx + 1),
            step.get('label', '未命名')[:30],
            step.get('request_method', ''),
            (step.get('request_url', '') or '')[:40],
            status_text,
            f"{step.get('duration', 0):.3f}s"
        ])

    step_table = Table(step_data, colWidths=[15*mm, 45*mm, 20*mm, 60*mm, 20*mm, 25*mm])
    step_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a9f82')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('WORDWRAP', (0, 0), (-1, -1), True),
    ]))
    elements.append(step_table)

    # 生成 PDF
    doc.build(elements)
    output.seek(0)
    return output.read()